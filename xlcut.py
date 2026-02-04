#!/usr/bin/env python3
"""XLCut - Convert XML files to Excel.

Drop XML files into the 'source' folder, run this script,
and find your Excel file in the 'output' folder.

Usage:
    python xlcut.py
"""

import sys
from pathlib import Path
from datetime import datetime
from collections import Counter

from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Folders
SCRIPT_DIR = Path(__file__).parent
SOURCE_DIR = SCRIPT_DIR / "source"
OUTPUT_DIR = SCRIPT_DIR / "output"


def parse_xml(file_path: Path) -> tuple[list[dict], list[str], str | None]:
    """Parse an XML file and return rows, columns, and the repeating element name."""
    with open(file_path, 'rb') as f:
        content = f.read()

    if not content.strip():
        return [], [], None

    root = etree.fromstring(content)

    # Find repeating elements
    repeating_tag = find_repeating_element(root)

    if repeating_tag is None:
        # Single element - treat as one row
        rows = [flatten_element(root)]
    else:
        elements = root.findall(f".//{repeating_tag}")
        rows = [flatten_element(elem) for elem in elements]

    # Extract columns
    columns = extract_columns(rows)

    return rows, columns, repeating_tag


def find_repeating_element(root) -> str | None:
    """Find the most likely repeating element (data rows)."""
    # Count direct children
    child_counts = Counter(child.tag for child in root)
    repeating = [(tag, count) for tag, count in child_counts.items() if count > 1]

    if repeating:
        return max(repeating, key=lambda x: x[1])[0]

    # Check one level deeper
    for child in root:
        grandchild_counts = Counter(gc.tag for gc in child)
        repeating = [(tag, count) for tag, count in grandchild_counts.items() if count > 1]
        if repeating:
            return max(repeating, key=lambda x: x[1])[0]

    return None


def flatten_element(element, prefix: str = "") -> dict:
    """Flatten an XML element into a dict with dot-notation keys."""
    result = {}

    # Add attributes
    for attr_name, attr_value in element.attrib.items():
        key = f"{prefix}@{attr_name}" if prefix else f"@{attr_name}"
        result[key] = attr_value

    children = list(element)

    if not children:
        # Leaf element
        text = element.text.strip() if element.text else ""
        if text and prefix:
            result[prefix.rstrip('.')] = text
        elif text:
            result[element.tag] = text
    else:
        # Has children
        for child in children:
            child_prefix = f"{prefix}{child.tag}." if prefix else f"{child.tag}."

            if len(list(child)) == 0 and not child.attrib:
                # Simple leaf
                text = child.text.strip() if child.text else ""
                key = f"{prefix}{child.tag}" if prefix else child.tag
                result[key] = text
            else:
                # Nested element
                child_result = flatten_element(child, child_prefix)
                result.update(child_result)

    return result


def extract_columns(rows: list[dict]) -> list[str]:
    """Extract unique column names preserving order."""
    seen = set()
    columns = []
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                columns.append(key)
    return columns


def write_sheet(ws, rows: list[dict], columns: list[str], include_source: bool = False):
    """Write rows to a worksheet with formatting."""
    # Add source file column if multiple files
    if include_source:
        columns = ["_source_file"] + [c for c in columns if c != "_source_file"]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row_data.get(col_name, "")

            # Try to convert numbers
            if value:
                try:
                    if '.' in str(value):
                        cell.value = float(value)
                    else:
                        cell.value = int(value)
                except (ValueError, TypeError):
                    cell.value = value
            else:
                cell.value = value

            cell.border = thin_border

            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")

    # Auto-fit columns (sample first 100 rows for performance)
    for col_idx, col_name in enumerate(columns, start=1):
        max_width = len(str(col_name))
        for row in rows[:100]:
            value = row.get(col_name, "")
            max_width = max(max_width, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 50)


def write_excel(all_rows: list[dict], columns: list[str], output_path: Path, source_files: list[str] = None):
    """Write rows to Excel with formatting, grouping by @type into separate sheets."""
    wb = Workbook()

    # Check if we have multiple source files
    include_source = source_files and len(set(source_files)) > 1

    # Group rows by @type attribute
    rows_by_type = {}
    for row in all_rows:
        row_type = row.get("@type", "data")
        if row_type not in rows_by_type:
            rows_by_type[row_type] = []
        rows_by_type[row_type].append(row)

    # If only one type, use the old behavior (single sheet)
    if len(rows_by_type) == 1:
        ws = wb.active
        ws.title = "Data"
        write_sheet(ws, all_rows, columns, include_source)
        wb.save(output_path)
        return {"Data": len(all_rows)}

    # Multiple types - create a sheet for each
    sheet_stats = {}
    first_sheet = True

    # Sort types for consistent ordering
    for row_type in sorted(rows_by_type.keys()):
        type_rows = rows_by_type[row_type]

        # Create sheet name (Excel limits to 31 chars, no special chars)
        sheet_name = str(row_type).replace("/", "-").replace("\\", "-")[:31]

        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Get columns that are actually used by this type (preserving order)
        type_columns = extract_columns(type_rows)
        # Remove @type from columns since it's redundant (it's the sheet name)
        type_columns = [c for c in type_columns if c != "@type"]

        write_sheet(ws, type_rows, type_columns, include_source)
        sheet_stats[sheet_name] = len(type_rows)

    wb.save(output_path)
    return sheet_stats


def extract_sales_items(file_path: Path) -> list[dict]:
    """Extract individual sold items from sale transactions."""
    with open(file_path, 'rb') as f:
        content = f.read()

    if not content.strip():
        return []

    root = etree.fromstring(content)
    items = []

    # Find all sale transactions
    for trans in root.findall(".//trans[@type='sale']"):
        # Extract header info
        header = trans.find("trHeader")
        if header is None:
            continue

        date_elem = header.find("date")
        date_val = date_elem.text if date_elem is not None else ""

        cashier_elem = header.find("cashier")
        cashier_name = cashier_elem.text if cashier_elem is not None else ""

        pos_elem = header.find("physicalRegisterID")
        register = pos_elem.text if pos_elem is not None else ""

        ticket_num = header.find("trTickNum")
        ticket_seq = ""
        if ticket_num is not None:
            seq_elem = ticket_num.find("trSeq")
            ticket_seq = seq_elem.text if seq_elem is not None else ""

        # Extract each line item
        tr_lines = trans.find("trLines")
        if tr_lines is None:
            continue

        for line in tr_lines.findall("trLine"):
            # Get the sign (1.00 = sale, -1.00 = payout/refund)
            sign_elem = line.find("trlSign")
            sign_val = sign_elem.text if sign_elem is not None else "1.00"

            try:
                sign = float(sign_val)
            except (ValueError, TypeError):
                sign = 1.0

            # Only include positive (sales) items
            if sign <= 0:
                continue

            # Get line total
            total_elem = line.find("trlLineTot")
            line_total = total_elem.text if total_elem is not None else "0.00"

            try:
                total_float = float(line_total)
            except (ValueError, TypeError):
                total_float = 0.0

            # Skip zero or negative totals
            if total_float <= 0:
                continue

            # Extract item details
            desc_elem = line.find("trlDesc")
            description = desc_elem.text if desc_elem is not None else ""

            qty_elem = line.find("trlQty")
            quantity = qty_elem.text if qty_elem is not None else "1"

            price_elem = line.find("trlUnitPrice")
            unit_price = price_elem.text if price_elem is not None else ""

            upc_elem = line.find("trlUPC")
            upc = upc_elem.text if upc_elem is not None else ""

            dept_elem = line.find("trlDept")
            department = dept_elem.text if dept_elem is not None else ""

            line_type = line.get("type", "")

            items.append({
                "Date": date_val,
                "Register": register,
                "Cashier": cashier_name,
                "Ticket#": ticket_seq,
                "Description": description,
                "Quantity": quantity,
                "Unit Price": unit_price,
                "Total": line_total,
                "Department": department,
                "UPC": upc,
                "Type": line_type,
            })

    return items


def write_items_sheet(ws, items: list[dict]):
    """Write sold items to a worksheet with nice formatting."""
    columns = ["Date", "Register", "Cashier", "Ticket#", "Description",
               "Quantity", "Unit Price", "Total", "Department", "UPC", "Type"]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Green for sales
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_idx, item in enumerate(items, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = item.get(col_name, "")

            # Convert numbers
            if col_name in ("Quantity", "Unit Price", "Total"):
                try:
                    cell.value = float(value) if value else 0.0
                    if col_name in ("Unit Price", "Total"):
                        cell.number_format = '$#,##0.00'
                except (ValueError, TypeError):
                    cell.value = value
            elif col_name == "Date" and value:
                # Format date nicely
                try:
                    dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
                    cell.value = dt.strftime("%Y-%m-%d %H:%M")
                except:
                    cell.value = value
            else:
                cell.value = value

            cell.border = thin_border

            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    # Auto-fit columns
    for col_idx, col_name in enumerate(columns, start=1):
        max_width = len(str(col_name))
        for item in items[:100]:
            value = item.get(col_name, "")
            max_width = max(max_width, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 50)


def main():
    print("=" * 50)
    print("XLCut - XML to Excel Converter")
    print("=" * 50)

    # Ensure folders exist
    SOURCE_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)

    # Find XML files in source folder
    xml_files = list(SOURCE_DIR.glob("*.xml"))

    if not xml_files:
        print(f"\nNo XML files found in: {SOURCE_DIR}")
        print(f"\nTo use:")
        print(f"  1. Put your XML files in: {SOURCE_DIR}")
        print(f"  2. Run this script again")
        print(f"  3. Find your Excel file in: {OUTPUT_DIR}")
        sys.exit(1)

    print(f"\nFound {len(xml_files)} XML file(s) in source folder")
    print("-" * 50)

    # Parse all files
    all_rows = []
    all_columns = set()
    column_order = []
    source_files = []
    all_items = []  # For sold items

    for xml_file in sorted(xml_files):
        try:
            rows, columns, repeating = parse_xml(xml_file)

            # Extract sold items
            items = extract_sales_items(xml_file)
            all_items.extend(items)

            if not rows:
                print(f"  {xml_file.name}: No data found, skipping")
                continue

            print(f"  {xml_file.name}: {len(rows)} transactions, {len(items)} items sold")

            # Track source file for each row
            for row in rows:
                row["_source_file"] = xml_file.name
                source_files.append(xml_file.name)

            all_rows.extend(rows)

            # Preserve column order
            for col in columns:
                if col not in all_columns:
                    all_columns.add(col)
                    column_order.append(col)

        except Exception as e:
            print(f"  {xml_file.name}: Error - {e}")

    if not all_rows and not all_items:
        print("\nError: No data extracted from any files")
        sys.exit(1)

    # Generate output filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"export_{timestamp}.xlsx"

    # Create workbook with Items Sold as first sheet
    wb = Workbook()

    if all_items:
        # First sheet: Items Sold (the main sheet user wants)
        ws_items = wb.active
        ws_items.title = "Items Sold"
        write_items_sheet(ws_items, all_items)

        # Calculate totals
        total_sales = sum(float(item.get("Total", 0)) for item in all_items)

        print("-" * 50)
        print(f"\n*** ITEMS SOLD: {len(all_items)} items, ${total_sales:,.2f} total ***")

    # Add transaction sheets if there's data
    if all_rows:
        # Group rows by @type attribute
        rows_by_type = {}
        for row in all_rows:
            row_type = row.get("@type", "data")
            if row_type not in rows_by_type:
                rows_by_type[row_type] = []
            rows_by_type[row_type].append(row)

        include_source = source_files and len(set(source_files)) > 1

        for row_type in sorted(rows_by_type.keys()):
            type_rows = rows_by_type[row_type]
            sheet_name = str(row_type).replace("/", "-").replace("\\", "-")[:31]

            ws = wb.create_sheet(title=sheet_name)
            type_columns = extract_columns(type_rows)
            type_columns = [c for c in type_columns if c != "@type"]

            write_sheet(ws, type_rows, type_columns, include_source)

    wb.save(output_path)

    print(f"\nOutput: {output_path}")
    print("\nDone!")


if __name__ == "__main__":
    main()
